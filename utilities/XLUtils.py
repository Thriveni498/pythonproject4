import openpyxl

def getRow(file,Sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheetname]
    return(sheet.max_row)

def getcolumn(file,Sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheetname]
    return(sheet.max_column)

def readdata(file,Sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheetname]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file, Sheetname, rownum, columnno, data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheetname]
    sheet.cell(row=rownum, column=columnno).value=data
    workbook.save(file)







